"""Offline grid compiler.

Reads the four supplied commission grids from ``data/raters/`` and emits one
normalized *rulepack* JSON per insurer into ``app/rulepacks/``. Each rulepack
retains the sheet + cell (or PDF page) provenance for every value the resolver
may read, so the runtime never has to open an XLSX/PDF and every rate can be
traced back to an exact cell.

Run:  python -m tools.compile_grids            (from backend/)
Check: python -m tools.compile_grids --check   (fails if output drifts)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path

import openpyxl

BACKEND = Path(__file__).resolve().parent.parent
RATERS = BACKEND.parent / "data" / "raters"
OUT = BACKEND / "app" / "rulepacks"

REL_XLSX = RATERS / "reliance-indusind/feb-2026/Reliance Broking Premier  FEB 26 Grid.xlsx"
GOD_XLSX = RATERS / "godigit/march-2026/Large Insurance Brokers Mar'26 - Shared.xlsx"
TATA_XLSX = RATERS / "tata-aig/march-2026/Tata AIG Standard Grid_Communication_Mar'26_F_v2_0212.xlsx"
HDFC_PDF = RATERS / "hdfc-ergo/feb-2025/Pvt Car New Grid Eff 1st Feb'25 (HDFC ergo) 1.pdf"


def _col_letter(idx0: int) -> str:
    """0-based column index -> Excel letter."""
    s = ""
    n = idx0 + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# --------------------------------------------------------------------------- #
# Reliance
# --------------------------------------------------------------------------- #
def compile_reliance() -> dict:
    wb = openpyxl.load_workbook(REL_XLSX, data_only=True)
    fname = REL_XLSX.name

    # RTO List: A Region_Code, C Region_City, D Market_Type, E Zone
    rto = wb["RTO List"]
    rto_map: dict[str, dict] = {}
    for row in rto.iter_rows(min_row=2):
        code = row[0].value
        if not code:
            continue
        key = str(code).strip().upper()
        rto_map[key] = {
            "region_city": (str(row[2].value).strip() if row[2].value else None),
            "market_type": (str(row[3].value).strip() if row[3].value else None),
            "zone": (str(row[4].value).strip().upper() if row[4].value else None),
            "cells": {
                "region_city": f"RTO List!C{row[0].row}",
                "zone": f"RTO List!E{row[0].row}",
            },
        }

    # PRIVATE CAR COMP, SAOD & STP
    sh = "PRIVATE CAR COMP, SAOD & STP"
    ws = wb[sh]
    pvt_rows = []
    current_zone = None
    for r in range(3, ws.max_row + 1):
        a, b, c, d, e, f = (ws.cell(r, i).value for i in range(1, 7))
        if a:
            current_zone = str(a).strip().upper()
        if b is None or c is None:
            continue
        pvt_rows.append(
            {
                "zone": current_zone,
                "rto_region": str(b).strip(),
                "petrol_bifuel_comp": c,
                "diesel_ev_comp": d,
                "sa_od": e,
                "stp": f,
                "row": r,
                "cells": {
                    "zone": f"{sh}!A{r}" if a else None,
                    "rto_region": f"{sh}!B{r}",
                    "petrol_bifuel_comp": f"{sh}!C{r}",
                    "diesel_ev_comp": f"{sh}!D{r}",
                    "sa_od": f"{sh}!E{r}",
                    "stp": f"{sh}!F{r}",
                },
            }
        )

    footnotes = []
    for r in range(2, 12):
        v = ws.cell(r, 8).value  # column H
        if v:
            footnotes.append({"text": str(v).strip(), "cell": f"{sh}!H{r}"})

    # Make/Model -> segment (kept for citation; the pvt-car COMP table is
    # zone x fuel only, but higher segments can matter for exclusions).
    mm = wb["Make Model wise segment"]
    seg_map = {}
    for row in mm.iter_rows(min_row=2):
        if row[0].value and row[1].value:
            seg_map[f"{str(row[0].value).strip().upper()}|{str(row[1].value).strip().upper()}"] = {
                "segment": str(row[2].value).strip() if row[2].value else None,
                "cell": f"Make Model wise segment!C{row[0].row}",
            }

    return {
        "insurer": "Reliance",
        "insurer_aliases": ["reliance", "reliance general", "reliance general insurance"],
        "file": fname,
        "effective": "Feb 2026 (table header: Pvt Car Old Business Grid - Jan-26)",
        "rate_sheet": sh,
        "rto_zone_map": rto_map,
        "pvt_car_rows": pvt_rows,
        "footnotes": footnotes,
        "make_model_segment": seg_map,
    }


# --------------------------------------------------------------------------- #
# Go Digit
# --------------------------------------------------------------------------- #
def compile_godigit() -> dict:
    wb = openpyxl.load_workbook(GOD_XLSX, data_only=True)
    fname = GOD_XLSX.name

    # 4W  RTO : B RTO Code, C 4WTP, D 4W Package, E 4WSAOD, F PLR Cluster
    ws = wb["4W  RTO"]
    rto_map = {}
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 2).value
        if not code:
            continue
        key = str(code).strip().upper().replace("-", "").replace(" ", "")
        rto_map[key] = {
            "tp_cluster": _s(ws.cell(r, 3).value),
            "package_cluster": _s(ws.cell(r, 4).value),
            "saod_cluster": _s(ws.cell(r, 5).value),
            "plr_cluster": _s(ws.cell(r, 6).value),
            "row": r,
            "cells": {
                "tp_cluster": f"4W  RTO!C{r}",
                "package_cluster": f"4W  RTO!D{r}",
                "saod_cluster": f"4W  RTO!E{r}",
            },
        }

    # 4W SATP : B Cluster, C Segment, D Age, E rate, F Note
    ws = wb["4W SATP"]
    satp_rows = []
    for r in range(3, ws.max_row + 1):
        cl = ws.cell(r, 2).value
        seg = ws.cell(r, 3).value
        if not cl or not seg:
            continue
        satp_rows.append(
            {
                "cluster": str(cl).strip(),
                "segment": str(seg).strip(),
                "age": _s(ws.cell(r, 4).value),
                "rate": ws.cell(r, 5).value,
                "note": _s(ws.cell(r, 6).value),
                "row": r,
                "cells": {"rate": f"4W SATP!E{r}", "note": f"4W SATP!F{r}"},
            }
        )

    # 4W New Business 1+3 : A Cluster, B..H make columns (row 3 header)
    ws = wb["4W New Business 1+3"]
    make_cols = {}
    for ci in range(2, ws.max_column + 1):
        h = ws.cell(3, ci).value
        if h:
            make_cols[str(h).strip()] = _col_letter(ci - 1)
    nb_rows = []
    for r in range(4, ws.max_row + 1):
        cl = ws.cell(r, 1).value
        if not cl:
            continue
        rates = {}
        for name, letter in make_cols.items():
            ci = openpyxl.utils.column_index_from_string(letter)
            rates[name] = {"value": ws.cell(r, ci).value, "cell": f"4W New Business 1+3!{letter}{r}"}
        nb_rows.append({"cluster": str(cl).strip(), "row": r, "rates": rates})

    return {
        "insurer": "Go Digit",
        "insurer_aliases": ["go digit", "godigit", "digit", "go digit general insurance"],
        "file": fname,
        "effective": "Mar 2026",
        "rto_cluster_map": rto_map,
        "satp_rows": satp_rows,
        "new_business_make_columns": list(make_cols.keys()),
        "new_business_rows": nb_rows,
        "notes": [
            "'4W SATP' is the stand-alone-TP grid for renewal / rollover business.",
            "'4W New Business 1+3' applies to brand-new (1+3 bundled) business only.",
            "Cluster 'All_India_Decline' and note 'declined' => segment not supported.",
        ],
    }


# --------------------------------------------------------------------------- #
# Tata AIG
# --------------------------------------------------------------------------- #
def compile_tataaig() -> dict:
    wb = openpyxl.load_workbook(TATA_XLSX, data_only=True)
    fname = TATA_XLSX.name
    ws = wb["Pvtcar"]

    # header row 5: F Concatenate, G Type, H Business Type, I Fuel, J Section,
    # K NCB, L Add On, M.. = cluster-city columns
    header_row = 5
    first_cluster_col = 13  # M
    cluster_columns = {}
    for ci in range(first_cluster_col, ws.max_column + 1):
        name = ws.cell(header_row, ci).value
        if name and str(name).strip():
            cluster_columns[str(name).strip()] = _col_letter(ci - 1)

    rows = []
    for r in range(6, ws.max_row + 1):
        seg = ws.cell(r, 7).value  # G
        biz = ws.cell(r, 8).value  # H
        fuel = ws.cell(r, 9).value  # I
        section = ws.cell(r, 10).value  # J
        if not seg or not section:
            continue
        rates = {}
        for name, letter in cluster_columns.items():
            ci = openpyxl.utils.column_index_from_string(letter)
            val = ws.cell(r, ci).value
            if val is not None:
                rates[name] = val
        rows.append(
            {
                "segment": str(seg).strip(),
                "business_type": _s(biz),
                "fuel": _s(fuel),
                "section": str(section).strip(),
                "ncb": _s(ws.cell(r, 11).value),
                "addon": _s(ws.cell(r, 12).value),
                "concat": _s(ws.cell(r, 6).value),
                "row": r,
                "rates": rates,
            }
        )

    guidelines = []
    gw = wb["General Guidelines"]
    for row in gw.iter_rows():
        for c in row:
            if isinstance(c.value, str) and len(c.value.strip()) > 12:
                guidelines.append({"text": c.value.strip(), "cell": f"General Guidelines!{c.coordinate}"})

    return {
        "insurer": "Tata AIG",
        "insurer_aliases": ["tata aig", "tata-aig", "tata aig general insurance"],
        "file": fname,
        "effective": "Mar 2026 (applicable from 1st Nov'25 policies)",
        "rate_sheet": "Pvtcar",
        "cluster_columns": cluster_columns,
        "pvtcar_rows": rows,
        "guidelines": guidelines,
        "notes": [
            "Package / SAOD % applies on OD premium; SATP % applies on Net premium.",
            "Visible 'Pvtcar' sheet holds the base commission; an additional GWP-slab "
            "incremental rule (portfolio level) is out of scope for a single-policy rate.",
        ],
    }


# --------------------------------------------------------------------------- #
# HDFC ERGO  (PDF grid -- hand-transcribed from the 2-page rate card, with
# page citations. Parsing the scanned tables directly is error-prone; the
# transcription is verified against the source image.)
# --------------------------------------------------------------------------- #
def compile_hdfc() -> dict:
    fname = HDFC_PDF.name
    # columns: package_petrol, package_nonpetrol_ncb, package_nonpetrol_nncb,
    #          saod_petrol,   saod_nonpetrol_ncb,    saod_nonpetrol_nncb
    zone1 = {
        "<10k":     [17.50, 17.50, 11.00, 17.50, 15.00, 11.00],
        "10k-50k":  [19.50, 19.50, 12.00, 19.50, 16.00, 12.00],
        "50k-1L":   [21.00, 21.00, 13.00, 21.00, 17.00, 13.00],
        "1L-2L":    [23.00, 23.00, 14.00, 23.00, 18.00, 14.00],
        ">2L":      [25.00, 25.00, 15.00, 25.00, 20.00, 15.00],
    }
    zone2 = {
        "<10k":     [15.00, 15.00, 10.00, 15.00, 11.00, 10.00],
        "10k-50k":  [17.50, 17.50, 10.00, 17.50, 12.00, 10.00],
        "50k-1L":   [20.00, 20.00, 12.00, 20.00, 13.00, 10.00],
        "1L-2L":    [22.00, 22.00, 12.00, 22.00, 14.00, 10.50],
        ">2L":      [23.00, 23.00, 12.50, 23.00, 15.00, 10.50],
    }
    cols = [
        "package_petrol", "package_nonpetrol_ncb", "package_nonpetrol_nncb",
        "saod_petrol", "saod_nonpetrol_ncb", "saod_nonpetrol_nncb",
    ]

    def table(z):
        return {slab: dict(zip(cols, vals)) for slab, vals in z.items()}

    # State -> zone mapping (page 1 & 2 of the rate card). Where a state is split
    # by city, list the Zone-1 regions explicitly; everything else in that state
    # falls to the state's default zone.
    state_zone = {
        "ASSAM": {"default": 1, "zone2_regions": ["Nagaon"], "page": 1},
        "BIHAR": {"default": 1, "page": 1},
        "JHARKHAND": {"default": 1, "page": 1},
        "WEST BENGAL": {"default": 1, "page": 1},
        "ARUNACHAL PRADESH": {"default": 1, "page": 1},
        "CHHATTISGARH": {"default": 2, "page": 1},
        "MANIPUR": {"default": 1, "page": 1},
        "MEGHALAYA": {"default": 1, "page": 1},
        "MIZORAM": {"default": 2, "page": 1},
        "NAGALAND": {"default": 1, "page": 1},
        "ODISHA": {"default": 1, "page": 1},
        "SIKKIM": {"default": 1, "page": 1},
        "TRIPURA": {"default": 1, "page": 1},
        "ANDHRA PRADESH": {"default": 1, "page": 1},
        "KARNATAKA": {"default": 2, "zone1_regions": ["Bangalore"], "page": 1},
        "TAMIL NADU": {"default": 2, "zone1_regions": ["Andamans"], "page": 1},
        "KERALA": {"default": 2, "page": 1},
        "TELANGANA": {"default": 1, "page": 1},
        "GOA": {"default": 1, "page": 1},
        "GUJARAT": {
            "default": 2,
            "zone1_regions": ["Ahmedabad", "Dadra & Nagar Haveli", "Daman", "Vadodara"],
            "page": 1,
        },
        "MAHARASHTRA": {"default": 1, "page": 1},
        "HARYANA": {"default": 2, "page": 2},
        "HIMACHAL PRADESH": {"default": 2, "page": 2},
        "PUNJAB": {"default": 2, "page": 2},
        "UTTAR PRADESH": {"default": 2, "page": 2},
        "UTTARAKHAND": {"default": 2, "page": 2},
        "JAMMU AND KASHMIR": {"default": 2, "page": 2},
        "DELHI": {"default": 1, "page": 2},
        "CHANDIGARH": {"default": 2, "page": 2},
        "MADHYA PRADESH": {"default": 2, "page": 2},
        "RAJASTHAN": {"default": 2, "page": 2},
    }

    return {
        "insurer": "HDFC ERGO",
        "insurer_aliases": ["hdfc ergo", "hdfc-ergo", "hdfc ergo general insurance"],
        "file": fname,
        "effective": "1st Feb'25",
        "source_kind": "pdf",
        "columns": cols,
        "slabs": ["<10k", "10k-50k", "50k-1L", "1L-2L", ">2L"],
        "slab_bounds": [
            ["<10k", 0, 10000],
            ["10k-50k", 10000, 50000],
            ["50k-1L", 50000, 100000],
            ["1L-2L", 100000, 200000],
            [">2L", 200000, None],
        ],
        "rate_tables": {"1": table(zone1), "2": table(zone2)},
        "rate_table_pages": {"1": "page 1, 'Zone 1' table", "2": "page 1, 'Zone 2' table"},
        "state_zone_map": state_zone,
        "state_zone_map_page": "page 1-2, 'Zone / State Name / Zone-1 / Zone-2' table",
        "footnotes": [
            {"text": "Slab will be calculated basis comprehensive + SAOD GWP on PVT CAR (SATP premium is excluded for slab achievement)", "page": 1},
            {"text": "*Non Petrol includes Diesel, CNG, LPG", "page": 1},
            {"text": "#New Business will be considered as NCB", "page": 1},
            {"text": "# EV and Hybrid is to be considered in Petrol Grid", "page": 1},
        ],
        "notes": [
            "Grid publishes OD commission only (Package & SAOD columns). No third-party "
            "commission column exists; TP payout is treated as 0%.",
        ],
    }


def _s(v):
    return str(v).strip() if v not in (None, "") else None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--check", action="store_true", help="fail if committed rulepacks differ")
    args = ap.parse_args()

    OUT.mkdir(parents=True, exist_ok=True)
    packs = {
        "reliance": compile_reliance(),
        "godigit": compile_godigit(),
        "tataaig": compile_tataaig(),
        "hdfc_ergo": compile_hdfc(),
    }
    drift = False
    for name, pack in packs.items():
        blob = json.dumps(pack, indent=2, ensure_ascii=False, sort_keys=True, default=str)
        path = OUT / f"{name}.json"
        if args.check:
            old = path.read_text() if path.exists() else ""
            if hashlib.sha256(old.encode()).hexdigest() != hashlib.sha256((blob + "\n").encode()).hexdigest():
                print(f"DRIFT: {path} differs from freshly compiled output")
                drift = True
        else:
            path.write_text(blob + "\n")
            print(f"wrote {path}  ({len(blob)} bytes, "
                  f"{_count(pack)} rate rows)")
    if args.check and drift:
        return 1
    return 0


def _count(pack: dict) -> int:
    for k in ("pvt_car_rows", "satp_rows", "pvtcar_rows"):
        if k in pack:
            return len(pack[k])
    return sum(len(t) for t in pack.get("rate_tables", {}).values())


if __name__ == "__main__":
    sys.exit(main())
